"""
consolidate_energy.py — Ireland Energy Dashboard
Reads raw Comtrade CSV/XLSX files and SEAI XLSX, produces clean processed CSVs.

Handles TWO different Comtrade file formats:
  Import CSVs:  year=refPeriodId, partner=partnerCode (ISO3), value=fobvalue
  Export XLSXs: year=refYear,     partner=partnerDesc (name), value=fobvalue

Raw files expected in data/raw/:
  comtrade_irl_energy_imports_2010_2021.csv
  comtrade_irl_energy_imports_2022_2025.csv
  comtrade_irl_energy_exports_2010_2021.xlsx
  comtrade_irl_energy_exports_2022_2025.xlsx
  seai_*.xlsx  (any file starting with seai)
"""
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from fx_rates import USD_TO_EUR, FALLBACK_RATE

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR  = os.path.join(THIS_DIR, "data", "raw")
PROC_DIR = os.path.join(THIS_DIR, "data", "processed")
os.makedirs(PROC_DIR, exist_ok=True)

# Country name canonicalization — same country appears under different names
# in raw UN Comtrade import vs export files. Map all variants to a single name.
COUNTRY_NORMALIZE = {
    'USA':               'United States',
    'Russia':            'Russian Federation',
    'Russian Fed.':      'Russian Federation',
    'Viet Nam':          'Vietnam',
    'Rep. of Korea':     'Republic of Korea',
    'South Korea':       'Republic of Korea',
    'China, Hong Kong SAR': 'Hong Kong',
    'Hong Kong, China':  'Hong Kong',
    'Taiwan Province of China': 'Taiwan',
    'Chinese Taipei':    'Taiwan',
    'Czech Republic':    'Czechia',
    'Turkey':            'Türkiye',
    'UK':                'United Kingdom',
    'Great Britain':     'United Kingdom',
    'Iran (Islamic Republic of)': 'Iran',
    'Bolivia (Plurinational State of)': 'Bolivia',
    'Venezuela (Bolivarian Rep. of)': 'Venezuela',
    'Dominican Rep.':    'Dominican Republic',
}

def normalize_country(name):
    """Map alternative spellings to canonical name."""
    if pd.isna(name): return name
    name = str(name).strip()
    return COUNTRY_NORMALIZE.get(name, name)

# ── ISO3 → Country name map ────────────────────────────────────────────────────
ISO3_MAP = {
    'GBR':'United Kingdom','NLD':'Netherlands','USA':'United States','NOR':'Norway',
    'ARE':'United Arab Emirates','RUS':'Russia','KWT':'Kuwait','AZE':'Azerbaijan',
    'QAT':'Qatar','DZA':'Algeria','NGA':'Nigeria','SAU':'Saudi Arabia','IRQ':'Iraq',
    'DEU':'Germany','FRA':'France','BEL':'Belgium','ESP':'Spain','DNK':'Denmark',
    'SWE':'Sweden','ITA':'Italy','POL':'Poland','FIN':'Finland','GRC':'Greece',
    'AUT':'Austria','CHE':'Switzerland','PRT':'Portugal','LUX':'Luxembourg',
    'IRL':'Ireland','LBY':'Libya','CAN':'Canada','AUS':'Australia','JPN':'Japan',
    'KOR':'Rep. of Korea','IND':'India','CHN':'China','SGP':'Singapore',
    'MYS':'Malaysia','BRA':'Brazil','MEX':'Mexico','ZAF':'South Africa',
    'EGY':'Egypt','MAR':'Morocco','TUR':'Türkiye','UKR':'Ukraine','BLR':'Belarus',
    'GEO':'Georgia','AGO':'Angola','GAB':'Gabon','NZL':'New Zealand',
    'IDN':'Indonesia','THA':'Thailand','VNM':'Vietnam','HKG':'Hong Kong',
    'BGR':'Bulgaria','CZE':'Czechia','HUN':'Hungary','ROU':'Romania',
    'SVK':'Slovakia','SVN':'Slovenia','EST':'Estonia','LVA':'Latvia',
    'LTU':'Lithuania','ISL':'Iceland','ALB':'Albania','SRB':'Serbia',
    'HRV':'Croatia','LKA':'Sri Lanka','BGD':'Bangladesh','ISR':'Israel',
    'PSE':'Palestine','BHR':'Bahrain','OMN':'Oman','JOR':'Jordan',
    'ECU':'Ecuador','PER':'Peru','CHL':'Chile','COL':'Colombia','ARG':'Argentina',
    'CIV':"Cote d'Ivoire",'SEN':'Senegal','UGA':'Uganda',
    'LBR':'Liberia','NCL':'New Caledonia','BMU':'Bermuda','BHS':'Bahamas',
    'TCA':'Turks and Caicos Islands','FRO':'Faroe Islands','GIB':'Gibraltar',
    'MLT':'Malta','CYP':'Cyprus','BRB':'Barbados','NIC':'Nicaragua',
    'BIH':'Bosnia and Herzegovina','GMB':'Gambia','MLI':'Mali',
    'UMI':'U.S. Minor Islands','BLM':'Saint Barthelemy',
}

# ── Continent map ──────────────────────────────────────────────────────────────
CONTINENT_MAP = {
    'United Kingdom':'Europe','Netherlands':'Europe','Belgium':'Europe',
    'Germany':'Europe','France':'Europe','Norway':'Europe','Sweden':'Europe',
    'Denmark':'Europe','Spain':'Europe','Italy':'Europe','Poland':'Europe',
    'Finland':'Europe','Greece':'Europe','Austria':'Europe','Switzerland':'Europe',
    'Portugal':'Europe','Luxembourg':'Europe','Ireland':'Europe',
    'Bulgaria':'Europe','Czechia':'Europe','Hungary':'Europe','Romania':'Europe',
    'Slovakia':'Europe','Slovenia':'Europe','Estonia':'Europe','Latvia':'Europe',
    'Lithuania':'Europe','Iceland':'Europe','Albania':'Europe','Serbia':'Europe',
    'Croatia':'Europe','Bosnia and Herzegovina':'Europe','Gibraltar':'Europe',
    'Malta':'Europe','Cyprus':'Europe','Faroe Islands':'Europe',
    'United States':'Americas','Canada':'Americas','Mexico':'Americas',
    'Brazil':'Americas','Argentina':'Americas','Colombia':'Americas',
    'Chile':'Americas','Peru':'Americas','Ecuador':'Americas',
    'Trinidad and Tobago':'Americas','Barbados':'Americas','Bahamas':'Americas',
    'Nicaragua':'Americas','Turks and Caicos Islands':'Americas',
    'Bermuda':'Americas','U.S. Minor Islands':'Americas',
    'Saudi Arabia':'Middle East','United Arab Emirates':'Middle East',
    'Kuwait':'Middle East','Iraq':'Middle East','Qatar':'Middle East',
    'Bahrain':'Middle East','Oman':'Middle East','Jordan':'Middle East',
    'Israel':'Middle East','Palestine':'Middle East','Yemen':'Middle East',
    'Russia':'Europe/Asia','Azerbaijan':'Europe/Asia','Ukraine':'Europe/Asia',
    'Belarus':'Europe/Asia','Georgia':'Europe/Asia',
    'China':'Asia','Japan':'Asia','Rep. of Korea':'Asia','India':'Asia',
    'Singapore':'Asia','Malaysia':'Asia','Indonesia':'Asia','Thailand':'Asia',
    'Vietnam':'Asia','Hong Kong':'Asia','Sri Lanka':'Asia','Bangladesh':'Asia',
    'Nigeria':'Africa','Algeria':'Africa','Libya':'Africa','Angola':'Africa',
    'Gabon':'Africa',"Cote d'Ivoire":'Africa','Egypt':'Africa',
    'Morocco':'Africa','Senegal':'Africa','Uganda':'Africa','Liberia':'Africa',
    'Mali':'Africa','South Africa':'Africa','Gambia':'Africa',
    'Australia':'Oceania','New Zealand':'Oceania','New Caledonia':'Oceania',
    'Türkiye':'Europe/Asia',
}

COORDS_MAP = {
    'United Kingdom':(55.4,-3.4),'Netherlands':(52.1,5.3),'Belgium':(50.5,4.5),
    'Germany':(51.2,10.5),'France':(46.2,2.2),'Norway':(60.5,8.5),
    'Sweden':(63.0,16.0),'Denmark':(56.0,10.0),'Spain':(40.0,-4.0),
    'Italy':(41.9,12.5),'Poland':(52.0,20.0),'United States':(39.5,-98.4),
    'Canada':(56.1,-106.3),'Saudi Arabia':(23.9,45.1),
    'United Arab Emirates':(23.4,53.8),'Kuwait':(29.3,47.5),
    'Qatar':(25.4,51.2),'Algeria':(28.0,1.7),'Nigeria':(9.1,8.7),
    'Russia':(61.5,105.0),'Azerbaijan':(40.1,47.6),
    'Libya':(26.3,17.2),'Iraq':(33.2,43.7),'China':(35.9,104.2),
    'Japan':(36.2,138.3),'Rep. of Korea':(35.9,127.8),'India':(20.6,78.9),
    'Singapore':(1.3,103.8),'Malaysia':(4.2,108.0),'Australia':(-25.3,133.8),
    'South Africa':(-30.6,22.9),'Brazil':(-14.2,-51.9),'Mexico':(23.6,-102.6),
    'Gibraltar':(36.1,-5.4),'Luxembourg':(49.6,6.1),
}

def map_product(v):
    v = str(v).lower()
    if 'crude' in v: return 'Crude petroleum oils'
    if any(x in v for x in ['refined','gasoline','diesel','kerosene','fuel oil',
                              'naphtha','gasoil','petroleum product','motor spirit',
                              'jet','aviation','not crude','preparations']):
        return 'Refined petroleum products'
    if any(x in v for x in ['natural gas','petroleum gas','lng','lpg',
                              'pipeline gas','gaseous hydrocarbon']):
        return 'Petroleum gas'
    if any(x in v for x in ['coal','coke','lignite','anthracite']):
        return 'Coal'
    if 'electri' in v:
        return 'Electrical energy'
    return str(v).strip().title()

JUNK_PARTNERS = {
    'world','areas, nes','other europe, nes','other asia, nes',
    'other africa, nes','other america, nes','other oceania, nes',
    'unspecified','total','not specified','free zones','bunkers',
    'low value trade','oecd','asean','opec','eu','european union',
    'special categories','other','unknown',
}

def to_float(series):
    cleaned = series.astype(str).str.strip().replace(
        {'N/A':'','n/a':'','-':'','NA':'','nan':'','None':''})
    return pd.to_numeric(cleaned, errors='coerce').fillna(0.0)

def add_geo(df):
    df['continent']   = df['partnerdesc'].map(lambda x: CONTINENT_MAP.get(x, 'Other'))
    df['partner_lat'] = df['partnerdesc'].map(lambda x: COORDS_MAP.get(x, (0,0))[0])
    df['partner_lon'] = df['partnerdesc'].map(lambda x: COORDS_MAP.get(x, (0,0))[1])
    return df

# ── Load import CSVs ───────────────────────────────────────────────────────────
print('\n--- Loading import CSV files ---')
import_frames = []

for key, fname in [
    ('imports_2010_2021', 'comtrade_irl_energy_imports_2010_2021.csv'),
    ('imports_2022_2025', 'comtrade_irl_energy_imports_2022_2025.csv'),
]:
    path = os.path.join(RAW_DIR, fname)
    if not os.path.exists(path):
        print(f'  MISSING: {fname}')
        continue
    print(f'  Loading: {fname}')
    try:
        df = pd.read_csv(path, low_memory=False, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(path, low_memory=False, encoding='latin-1')

    # Year
    df['refyear'] = pd.to_numeric(df['refPeriodId'], errors='coerce').fillna(0).astype(int)

    # Partner
    df['partnerdesc'] = df['partnerCode'].astype(str).map(
        lambda c: ISO3_MAP.get(c.strip(), c.strip()))
    df = df[df['partnerCode'].astype(str).str.strip() != 'W00'].copy()
    df = df[~df['partnerdesc'].str.lower().isin(JUNK_PARTNERS)].copy()
    # Canonicalize country names so USA/United States, Russia/Russian Fed., etc.
    # become a single partner in downstream aggregations.
    df['partnerdesc'] = df['partnerdesc'].map(normalize_country)

    df['flowcode'] = 'M'

    # Product: detect if cmdCode column contains text descriptions (swapped columns)
    first_cmd = str(df['cmdCode'].dropna().iloc[0]) if len(df['cmdCode'].dropna()) > 0 else ''
    if len(first_cmd) > 10:
        df['cmddesc'] = df['cmdCode'].map(map_product)
        df['cmdcode'] = 0
    else:
        df['cmddesc'] = df['cmdDesc'].map(map_product)
        df['cmdcode'] = pd.to_numeric(df['cmdCode'], errors='coerce').fillna(0).astype(int)

    # Value — UN Comtrade is in USD; convert to EUR using year-specific rate
    df['primaryvalue'] = to_float(df['fobvalue'])
    fx_series = df['refyear'].map(USD_TO_EUR).fillna(FALLBACK_RATE)
    df['primaryvalue'] = df['primaryvalue'] * fx_series  # now in EUR

    df = df[(df['refyear'] >= 1990) & (df['refyear'] <= 2030)].copy()
    df = add_geo(df)
    df['_source'] = key

    keep = ['refyear','flowcode','partnerdesc','cmdcode','cmddesc',
            'primaryvalue','continent','partner_lat','partner_lon','_source']
    df = df[keep].copy()

    imp_val = df['primaryvalue'].sum()
    print(f'    → M | {df["partnerdesc"].nunique()} partners | '
          f'years {sorted(df["refyear"].unique().tolist())[:2]}..{df["refyear"].max()} | '
          f'€{imp_val/1e9:.2f}bn')
    import_frames.append(df)

# ── Load export XLSXs ─────────────────────────────────────────────────────────
print('\n--- Loading export XLSX files ---')
export_frames = []

for key, fname in [
    ('exports_2010_2021', 'comtrade_irl_energy_exports_2010_2021.xlsx'),
    ('exports_2022_2025', 'comtrade_irl_energy_exports_2022_2025.xlsx'),
]:
    path = os.path.join(RAW_DIR, fname)
    if not os.path.exists(path):
        print(f'  MISSING: {fname}')
        continue
    print(f'  Loading: {fname}')
    df = pd.read_excel(path)

    df['refyear'] = pd.to_numeric(df['refYear'], errors='coerce').fillna(0).astype(int)

    df['partnerdesc'] = df['partnerDesc'].astype(str).str.strip()
    df = df[~df['partnerdesc'].str.lower().isin(JUNK_PARTNERS)].copy()
    df = df[df['partnerdesc'] != '0'].copy()
    # Canonicalize country names (same as imports path)
    df['partnerdesc'] = df['partnerdesc'].map(normalize_country)

    df['flowcode'] = 'X'

    df['cmddesc']  = df['cmdDesc'].map(map_product)
    df['cmdcode']  = pd.to_numeric(df['cmdCode'], errors='coerce').fillna(0).astype(int)

    # UN Comtrade is in USD; convert to EUR using year-specific rate
    df['primaryvalue'] = to_float(df['fobvalue'])
    fx_series = df['refyear'].map(USD_TO_EUR).fillna(FALLBACK_RATE)
    df['primaryvalue'] = df['primaryvalue'] * fx_series  # now in EUR
    df = df[df['primaryvalue'] > 0].copy()

    df = df[(df['refyear'] >= 1990) & (df['refyear'] <= 2030)].copy()
    df = add_geo(df)
    df['_source'] = key

    keep = ['refyear','flowcode','partnerdesc','cmdcode','cmddesc',
            'primaryvalue','continent','partner_lat','partner_lon','_source']
    df = df[keep].copy()

    exp_val = df['primaryvalue'].sum()
    print(f'    → X | {df["partnerdesc"].nunique()} partners | '
          f'years {sorted(df["refyear"].unique().tolist())[:2]}..{df["refyear"].max()} | '
          f'€{exp_val/1e9:.2f}bn')
    export_frames.append(df)

# ── Combine ────────────────────────────────────────────────────────────────────
all_frames = import_frames + export_frames
if not all_frames:
    print('ERROR: No data loaded. Check raw file paths.')
    exit(1)

combined = pd.concat(all_frames, ignore_index=True)
combined['refyear'] = combined['refyear'].astype(int)

print(f'\n--- Summary ---')
print(f'  Combined shape: {combined.shape}')
print(f'  Flowcodes: {combined["flowcode"].value_counts().to_dict()}')
print(f'  Years: {sorted(combined["refyear"].unique().tolist())}')
print(f'  Import partners: {combined[combined["flowcode"]=="M"]["partnerdesc"].nunique()}')
print(f'  Export partners: {combined[combined["flowcode"]=="X"]["partnerdesc"].nunique()}')
print(f'  Products: {combined["cmddesc"].unique().tolist()}')
print(f'  Total value: €{combined["primaryvalue"].sum()/1e9:.1f}bn')

out_ct = os.path.join(PROC_DIR, 'comtrade_consolidated.csv')
combined.to_csv(out_ct, index=False)
print(f'\n✓ Saved Comtrade → {out_ct}')

# ── Load SEAI ──────────────────────────────────────────────────────────────────
seai_candidates = [f for f in os.listdir(RAW_DIR) if f.lower().startswith('seai')]
if seai_candidates:
    seai_path = os.path.join(RAW_DIR, seai_candidates[0])
    print(f'\n--- Loading SEAI: {seai_candidates[0]} ---')
    FUEL_COLS = [
        'coal','bituminous_coal','anthracite_manufactured_ovoids','coke',
        'lignite_brown_coal_briquettes','peat','milled_peat','sod_peat','briquettes',
        'oil','crude','refinery_feedstocks','refinery_gas','gasoline','kerosene',
        'jet_kerosene','fueloil','lpg','gasoil_diesel_derv','petroleum_coke','naphta',
        'bitumen','white_spirit','lubricants','natural_gas','renewables','hydro',
        'wind','biomass','renewable_waste','landfill_gas','biogas','biodiesel',
        'bioethanol','solar_photovoltaic','solar_thermal','ambient_heat',
        'non_renewable_waste','electricity','heat','total'
    ]
    wb = load_workbook(seai_path, read_only=True, data_only=True)
    seai_rows = []
    for sheet_name in wb.sheetnames:
        try: year = int(sheet_name.strip())
        except ValueError: continue
        ws = wb[sheet_name]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0] or str(row[0]).startswith('='): continue
            flow = str(row[0]).strip()
            if not flow: continue
            rec = {'year': year, 'flow': flow}
            for i, f in enumerate(FUEL_COLS):
                v = row[i+2] if i+2 < len(row) else None
                if isinstance(v, str) and v.startswith('='): v = None
                rec[f] = v
            seai_rows.append(rec)
    wb.close()
    seai_df = pd.DataFrame(seai_rows)
    for c in FUEL_COLS:
        seai_df[c] = pd.to_numeric(seai_df[c], errors='coerce')
    seai_out = os.path.join(PROC_DIR, 'seai_energy_balance.csv')
    seai_df.to_csv(seai_out, index=False)
    print(f'  SEAI shape: {seai_df.shape}')
    print(f'  Years: {seai_df["year"].min()}–{seai_df["year"].max()}')
    print(f'  Saved → {seai_out}')
else:
    print('\n  No SEAI file found (file must start with "seai")')

print(f'\n✓ All done. Outputs in: {PROC_DIR}')
